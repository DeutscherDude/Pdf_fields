from fileinput import filename
from typing import List
import openpyxl
from openpyxl.reader.excel import load_workbook


def insert_new_row(values: List[str], file_name: str) -> None:
    wb = load_workbook(filename = file_name)
    ws = wb['Sheet1']
    row = ws.max_row + 1

    for col, entry in enumerate(values, start = 1):
        ws.cell(row = row, column = col, value = entry)

    wb.save(file_name)

def insert_to_cells(values: List[str], file_name: str, cells: List[str]) -> None:
    wb = load_workbook(filename= file_name)
    ws = wb['Sheet2']

    cell_c = 0
    for val in values:
        ws[cells[cell_c]] = val
        ws['A1'] = 'Bitch'
        cell_c += 1

    wb.save(file_name)
